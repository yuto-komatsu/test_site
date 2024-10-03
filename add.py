import streamlit as st
import datetime
from datetime import date
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import jpholiday
from openpyxl.styles import Border, Side, Font
from openpyxl.styles.alignment import Alignment

border_topthick = Border(top=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_bottomthick = Border(bottom=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_sidethick = Border(left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_topleft = Border(top=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
)

border_topcenter = Border(top=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_topright = Border(top=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                bottom=Side(style='thin', color='000000')
)

border_bottomleft = Border(bottom=Side(style='thick', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_bottomright = Border(bottom=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000')
)

border_bottomcenter = Border(bottom=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_left = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thick', color='000000'),
                right=Side(style='thin', color='000000')
)

border_right = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000')
)

border_allthin = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
)


band_list = {}
week = {}
st.session_state["page_control"] = 0


def change_page():
  # ページ切り替えボタンコールバック
  st.session_state["page_control"] += 1

def band_list_making():
  i = 1
  while st.session_state["sheet"].cell(row=5 + i, column=2).value is not None:
      band_list[i] = st.session_state["sheet"].cell(row=5 + i, column=2).value
      i += 1
  band_sum = len(band_list)
  return band_sum

def option_select():
    max_practice = st.selectbox(
        '最大練習回数',
        [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        index=0,
        placeholder="練習回数を選択してください"
    )
    return max_practice

def kinshi_select():
  # 日付を保存するためのセッション状態を初期化
  if 'dates_list' not in st.session_state:
    st.session_state['dates_list'] = []

  # 日付入力
  selected_date = st.date_input('日付を選択してください:', date.today())


  # 日付を追加するボタン
  if st.button('日付を追加'):
    if selected_date < st.session_state["start_day"] or selected_date > st.session_state["end_day"]:
      st.error('シフト期間内の日付を入力してください。')
      st.stop()
    st.session_state['dates_list'].append(selected_date)
    st.success(f'{selected_date} が追加されました。')

  if st.button('リセット'):
    st.session_state['dates_list'] = []
    st.success('リセットされました。')

  # 追加された日付と今日の日付の差分を辞書に登録（キーを1, 2, 3, ...とする）
  st.session_state["kinshi"] = {int(i + 1): (d - st.session_state["start_day"] + datetime.timedelta(days=1)).days for i, d in enumerate(st.session_state['dates_list'])}

  # 現在の追加済みの日付を表示
  st.write('追加された日付一覧:')
  i = 0
  while i < len(st.session_state['dates_list']):
    st.write(st.session_state['dates_list'][i])
    i += 1


  # 日付の差分を含む辞書を表示
  st.write('日付と今日との差分（日数）の辞書:', st.session_state["kinshi"])

def week_judge(start_day, vacation):
  # 平日と土日祝の判別
  current_day = start_day
  for i in range(1, day_sum + 1):
    if not vacation:
      if jpholiday.is_holiday(current_day) or current_day.weekday() in [5, 6]:
        week[i] = 1  # 祝日または土日
      else:
        week[i] = 0  # 平日
    else:
      week[i] = 1  # 長期休暇中
    current_day += datetime.timedelta(days=1)



def input_date():
  book1 = openpyxl.Workbook()
  for i in range(1, band_sum + 1):
    book1.create_sheet(index=0, title=band_list[i])
    sheet = book1[band_list[i]]
    for t in range(1, 8):
        sheet.cell(row=2 + t, column=2).value = str(t) + "限"
    calc_day = st.session_state["start_day"]
    j = 1
    while calc_day <= st.session_state["end_day"]:
        sheet.cell(row=2, column=2 + j).value = str(calc_day.month) + "/" + str(calc_day.day)
        j += 1
        calc_day += datetime.timedelta(days=1)

    #枠線作成
    #B列目
    sheet.cell(row=2, column=2).border = border_topleft
    for t in range(1,8):
      sheet.cell(row=2+t, column=2).border = border_left
    sheet.cell(row=9, column=2).border = border_bottomleft
    #2行目
    for j in range(1, day_sum):
      sheet.cell(row=2, column=2+j).border = border_topcenter
    #右端
    sheet.cell(row=2, column=day_sum+2).border = border_topright
    for t in range(1,8):
      sheet.cell(row=2+t, column=day_sum+2).border = border_right
    sheet.cell(row=9, column=day_sum+2).border = border_bottomright
    #中
    for i in range(1, day_sum):
      for t in range(1,7):
        sheet.cell(row=2+t, column=2+i).border = border_allthin
    #下
    for j in range(1, day_sum):
      sheet.cell(row=9, column=2+j).border = border_bottomcenter

    #書式設定
    font = Font(name="游ゴシック",size=11,bold=True)
    for t in range(1, 9):
      for j in range(1, day_sum+2):
        sheet.cell(row=1+t, column=1+j).font = font
        sheet.cell(row=1+t, column=1+j).alignment = Alignment(horizontal = 'left', vertical = 'center')


    # デフォルトで作成されるシートを削除
  if 'Sheet' in book1.sheetnames:
      book1.remove(book1['Sheet'])

    # バイトストリームにExcelファイルを保存
  buffer = BytesIO()
  book1.save(buffer)
  buffer.seek(0)

  # StreamlitのダウンロードボタンでExcelファイルをダウンロード
  st.download_button(
      label="ダウンロード",
      data=buffer,
      file_name='downloaded_file2.xlsx',
      mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
  )




# Webアプリのタイトル
st.title('シフトスケジュール最適化')


# #ページ１：参加バンド登録
# uploaded_file_path = 'シフト希望表.xlsx'
# # ファイルをバイトとして読み込む
# with open(uploaded_file_path, 'rb') as file:
#     band_listfile = file.read()

if st.session_state["page_control"] == 0:
  st.header('１．参加バンドの登録')
  st.caption('ダウンロードボタンからテンプレートをダウンロードして、出演バンドを記入してください。')
  st.caption('記入を終えたファイルをアップロードしてください。')

  # st.download_button(
  #     label="テンプレートをダウンロード",
  #     data=band_listfile,
  #     file_name='downloaded_file.xlsx',
  #     mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
  # )

  st.session_state["uploaded_file1"] = st.file_uploader("バンド名簿をアップロード", type=["xlsx"],key = "バンド名簿")
  if st.session_state["uploaded_file1"] is not None:
      change_page()



#ページ２：ライブ情報の入力
if "page_control" in st.session_state and st.session_state["page_control"] == 1:
    # st.session_state['uploaded'] = True
    st.header('２．ライブ情報の入力')
    st.session_state["book"] = load_workbook(st.session_state["uploaded_file1"])
    st.session_state["sheet"] = st.session_state["book"]["概要"]
    band_sum = band_list_making()

    st.session_state["start_day"] = st.date_input('シフト開始日を入力してください。', datetime.date(2024, 8, 22))
    st.session_state["end_day"] = st.date_input('シフト終了日を入力してください。', datetime.date(2024, 9, 9))
    if st.session_state["start_day"] > st.session_state["end_day"]:
        st.error('開始日は終了日より後の日付を入力してください。')
        st.stop()

    day_sum = (st.session_state["end_day"] - st.session_state["start_day"] + datetime.timedelta(days=1)).days
    max_practice = option_select()
    vacation = st.toggle("長期休暇期間")
    d = st.toggle("部室利用禁止日あり")
    if d == True:
      kinshi_select()
    week_judge(st.session_state["start_day"], vacation)

    if st.button("入力完了"):
        change_page()

#ページ３：希望日入力
if "page_control" in st.session_state and st.session_state["page_control"] == 2:
  st.header('３．練習希望日時の入力')
  input_date()
  st.write("記入を終えたファイルをアップロードしてください。")

  st.session_state["kibou_file"] = st.file_uploader("シフト希望表をアップロード", type=["xlsx"],key = "希望")
  if st.session_state["kibou_file"] is not None:
    # change_page()
    st.session_state["page_control"] = 3


#ページ４：最適化の実行
if "page_control" in st.session_state and st.session_state["page_control"] == 3:
  st.write('４．最適化の実行')
  st.stop()

#         # セッションからファイルを読み込む
#         kibou_file = st.session_state['kibou_file']
#         book = load_workbook(kibou_file)
#         st.success('ファイルが正常に読み込まれました。')
#         # 最適化の処理をここに追加
#     except Exception as e:
#         st.error(f'ファイルの読み込みに失敗しました: {e}')
